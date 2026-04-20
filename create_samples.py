"""
Create sample Excel files for testing BulkMailer Pro.
Run: python create_samples.py
"""
import openpyxl
from pathlib import Path

def create_personalized_sample():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recipients"
    
    # Headers
    ws.append(["Name", "Email"])
    
    # Sample data
    samples = [
        ("Alice Johnson", "alice@example.com"),
        ("Bob Smith", "bob@example.com"),
        ("Carol Williams", "carol@example.com"),
        ("David Brown", "david@example.com"),
        ("Emma Davis", "emma@example.com"),
    ]
    for row in samples:
        ws.append(row)
    
    # Style headers
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")
    
    path = Path("sample_personalized.xlsx")
    wb.save(path)
    print(f"Created: {path}")


def create_bulk_sample():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"
    
    emails = [
        ("user1@example.com",),
        ("user2@example.com",),
        ("user3@example.com",),
        ("user4@example.com",),
        ("user5@example.com",),
    ]
    for row in emails:
        ws.append(row)
    
    path = Path("sample_bulk.xlsx")
    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    create_personalized_sample()
    create_bulk_sample()
    print("\nSample files ready! Use them in BulkMailer Pro.")
